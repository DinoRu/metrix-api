# app/workers/tasks/meter_tasks.py
from __future__ import annotations
import base64, io, logging
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone

from sqlalchemy.orm import Session
from sqlalchemy import select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert

from openpyxl import load_workbook

from app.core.celery_app import celery_app
from app.database import AsyncSessionLocal
from app.models.meter import Meter
from app.models.task import TaskResult, TaskStatus

logger = logging.getLogger(__name__)


@celery_app.task(bind=True, name="app.workers.tasks.meter_tasks.import_meters_from_file")
def import_meters_from_file(
    self,
    *,
    file_content_b64: str,
    file_name: str,
    user_id: str,
    file_type: str = "xlsx",
    task_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Import des compteurs depuis un fichier Excel (.xlsx/.xls) en tâche de fond.
    - Insertions par lots avec ON CONFLICT DO NOTHING (sur meter_number)
    - Progression: TaskResult.progress + Celery update_state
    - Rollback propre en cas d'erreur
    """
    if file_type.lower() not in ("xlsx", "xls"):
        raise ValueError("Le fichier doit être .xlsx ou .xls")

    def safe_progress(db: Session, tid: str, current: int, total: int, success: int, failed: int):
        """Met à jour la progression en protégeant la session (rollback si nécessaire)."""
        percent = int(current * 100 / max(total, 1)) if total else 0
        try:
            # Celery state
            self.update_state(
                state="PROCESSING",
                meta={"current": current, "total": total, "success": success, "failed": failed, "percent": percent},
            )
            # BD
            db.execute(
                update(TaskResult)
                .where(TaskResult.id == tid)
                .values(
                    status=TaskStatus.PROCESSING,
                    progress={
                        "current": current,
                        "total": total,
                        "success": success,
                        "failed": failed,
                        "percent": percent,
                    },
                )
            )
            db.commit()
        except Exception:
            db.rollback()
            try:
                db.execute(
                    update(TaskResult)
                    .where(TaskResult.id == tid)
                    .values(
                        status=TaskStatus.PROCESSING,
                        progress={
                            "current": current,
                            "total": total,
                            "success": success,
                            "failed": failed,
                            "percent": percent,
                        },
                    )
                )
                db.commit()
            except Exception:
                db.rollback()  # on abandonne la MAJ de progression si la DB est KO

    tid = task_id or self.request.id
    db = AsyncSessionLocal
    try:
        # Assure un enregistrement TaskResult (si non créé côté API)
        tr = db.execute(select(TaskResult).where(TaskResult.id == tid)).scalar_one_or_none()
        if not tr:
            tr = TaskResult(
                id=tid,
                task_name="import_meters_from_file",
                user_id=user_id,
                status=TaskStatus.PENDING,
                params={"file_name": file_name, "ext": file_type},
                progress={"current": 0, "total": 0, "success": 0, "failed": 0, "percent": 0},
                started_at=datetime.now(timezone.utc),
            )
            db.add(tr)
            db.commit()

        # ===== Décode le payload =====
        raw = base64.b64decode(file_content_b64)
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = wb.active

        # Ligne 2 = en-têtes russes
        RUS_COLS = {
            "id_code": "Идентификационный код",
            "address": "Адрес",
            "client_name": "Наименование объекта сети",
            "meter_type": "Тип прибора учета",
            "meter_number": "Номер ПУ",
            "prev_reading": "Предыдущие показания",
            "curr_reading": "Текущие показания",
        }
        REQUIRED = [RUS_COLS["id_code"], RUS_COLS["meter_type"], RUS_COLS["meter_number"]]

        headers_row2 = [str(c.value).strip() if c.value else None for c in sheet[2]]
        header_index = {h: i for i, h in enumerate(headers_row2) if h}
        missing = [c for c in REQUIRED if c not in header_index]
        if missing:
            raise ValueError(f"Colonnes manquantes: {missing}. Colonnes détectées (ligne 2): {headers_row2}")

        total_rows = max(0, sheet.max_row - 2)
        safe_progress(db, tid, 0, total_rows, 0, 0)

        success = failed = processed = 0
        errors: List[dict] = []

        # Batching & throttle
        BATCH = 500
        PROGRESS_EVERY = 50  # MAJ progression toutes les 50 lignes pour limiter les commits
        buffer: List[dict] = []

        # Déduplication fichier (évite doublons à l’intérieur du même fichier)
        seen_numbers = set()

        def get_val(row, colname):
            idx = header_index.get(colname)
            return row[idx] if idx is not None and idx < len(row) else None

        def flush():
            """Insert bulk avec ON CONFLICT DO NOTHING sur meter_number."""
            nonlocal buffer, success, failed
            if not buffer:
                return
            try:
                stmt = insert(Meter.__table__).values(buffer)
                # 🔑 Si ta contrainte unique porte un autre nom/colonnes, adapte ici :
                stmt = stmt.on_conflict_do_nothing(index_elements=["meter_number"])
                result = db.execute(stmt)  # result.rowcount peut être -1 selon le driver
                db.commit()
            except IntegrityError as e:
                # Très rare ici (on_conflict), mais on sécurise
                logger.warning("Bulk insert integrity error, fallback piece-by-piece: %s", e)
                db.rollback()
                for row in buffer:
                    try:
                        db.execute(
                            insert(Meter.__table__).values(row).on_conflict_do_nothing(index_elements=["meter_number"])
                        )
                        db.commit()
                    except Exception as ex:
                        db.rollback()
                        errors.append({
                            "row": None,
                            "meter_number": row.get("meter_number"),
                            "meter_id_code": row.get("meter_id_code"),
                            "error": str(ex),
                        })
                        failed += 1
            except Exception as e:
                db.rollback()
                logger.exception("Bulk insert failed: %s", e)
                # on tente pièce par pièce
                for row in buffer:
                    try:
                        db.execute(
                            insert(Meter.__table__).values(row).on_conflict_do_nothing(index_elements=["meter_number"])
                        )
                        db.commit()
                    except Exception as ex:
                        db.rollback()
                        errors.append({
                            "row": None,
                            "meter_number": row.get("meter_number"),
                            "meter_id_code": row.get("meter_id_code"),
                            "error": str(ex),
                        })
                        failed += 1
            finally:
                buffer.clear()

        # Parcours des lignes 3..N
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            processed += 1
            try:
                meter_id_code = (str(get_val(row, RUS_COLS["id_code"])).strip()
                                 if get_val(row, RUS_COLS["id_code"]) else None)
                meter_number = (str(get_val(row, RUS_COLS["meter_number"])).strip()
                                if get_val(row, RUS_COLS["meter_number"]) else None)
                meter_type = (str(get_val(row, RUS_COLS["meter_type"])).strip()
                              if get_val(row, RUS_COLS["meter_type"]) else None)

                if not meter_id_code or not meter_number or not meter_type:
                    failed += 1
                    errors.append({"row": row_idx, "error": "Champs requis manquants (id/num/type)."})
                    if processed % PROGRESS_EVERY == 0:
                        safe_progress(db, tid, processed, total_rows, success, failed)
                    continue

                # Déduplication intra-fichier
                if meter_number in seen_numbers:
                    failed += 1
                    errors.append({"row": row_idx, "meter_number": meter_number, "error": "Duplicate in file"})
                    if processed % PROGRESS_EVERY == 0:
                        safe_progress(db, tid, processed, total_rows, success, failed)
                    continue
                seen_numbers.add(meter_number)

                # Préparer la ligne pour bulk insert
                row_dict = {
                    "meter_id_code": meter_id_code,
                    "meter_number": meter_number,
                    "type": meter_type,
                    # colonnes optionnelles si présentes dans ton modèle
                    "location_address": (str(get_val(row, RUS_COLS["address"])).strip()
                                         if get_val(row, RUS_COLS["address"]) else None),
                    "client_name": (str(get_val(row, RUS_COLS["client_name"])).strip()
                                    if get_val(row, RUS_COLS["client_name"]) else None),
                    "prev_reading_value": None,  # adapter si tu veux parser "Предыдущие показания"
                    "last_reading_date": None,   # idem si tu as une date
                    "status": "active",
                    "meter_metadata": {},
                }
                buffer.append(row_dict)
                success += 1

                if len(buffer) >= BATCH:
                    flush()

            except Exception as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e)})

            # Progression throttlée
            if processed % PROGRESS_EVERY == 0:
                safe_progress(db, tid, processed, total_rows, success, failed)

        # Flush final + progression finale
        flush()
        safe_progress(db, tid, processed, total_rows, success, failed)

        result = {
            "file": file_name,
            "success": success,
            "failed": failed,
            "total": total_rows,
            "errors": errors[:200],  # limite le retour
        }

        # Clôture TaskResult
        try:
            db.execute(
                update(TaskResult)
                .where(TaskResult.id == tid)
                .values(
                    status=TaskStatus.COMPLETED,
                    result=result,
                    completed_at=datetime.now(timezone.utc),
                    progress={"percent": 100, "current": processed, "total": total_rows,
                              "success": success, "failed": failed},
                )
            )
            db.commit()
        except Exception:
            db.rollback()
            db.execute(
                update(TaskResult)
                .where(TaskResult.id == tid)
                .values(
                    status=TaskStatus.COMPLETED,
                    result=result,
                    completed_at=datetime.now(timezone.utc),
                )
            )
            db.commit()

        return {"task_id": tid, "status": "completed", **result}

    except Exception as e:
        logger.exception(f"[{tid}] Import meters failed: {e}")
        db.rollback()
        try:
            db.execute(
                update(TaskResult)
                .where(TaskResult.id == tid)
                .values(
                    status=TaskStatus.FAILED,
                    error_message=str(e),
                    completed_at=datetime.now(timezone.utc),
                )
            )
            db.commit()
        except Exception:
            db.rollback()
        raise
    finally:
        db.close()


@celery_app.task(name="app.workers.tasks.meter_tasks.ping")
def ping():
    return "pong"
