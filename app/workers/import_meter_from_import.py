# app/services/meter_import.py
from __future__ import annotations
import base64, io, logging
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone

from sqlalchemy.orm import Session
from sqlalchemy import select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert

from openpyxl import load_workbook

from app.database import SessionLocalSync
from app.models.meter import Meter
from app.models.task import TaskResult, TaskStatus

logger = logging.getLogger(__name__)


def import_meters_from_file(
    *,
    file_content_b64: str,
    file_name: str,
    user_id: str,
    file_type: str = "xlsx",
    task_id: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Import synchrone de compteurs depuis un fichier Excel (.xlsx/.xls).
    - Seul id_code est obligatoire
    - Insertions par lots avec ON CONFLICT DO NOTHING (sur meter_number)
    - Mise à jour en base de l’état de la tâche
    """
    if file_type.lower() not in ("xlsx", "xls"):
        raise ValueError("Le fichier doit être .xlsx ou .xls")

    def safe_progress(db: Session, tid: str, current: int, total: int, success: int, failed: int):
        """Met à jour la progression (rollback si problème)."""
        percent = int(current * 100 / max(total, 1)) if total else 0
        try:
            db.execute(
                update(TaskResult)
                .where(TaskResult.id == tid)
                .values(
                    status=TaskStatus.PROCESSING,
                    progress={
                        "current": current,
                        "total": total,
                        "success": success,
                        "failed": failed,
                        "percent": percent,
                    },
                )
            )
            db.commit()
        except Exception:
            db.rollback()

    tid = task_id or f"manual-{datetime.utcnow().timestamp()}"
    db = SessionLocalSync()
    try:
        # Assurer un enregistrement TaskResult (si non existant)
        tr = db.execute(select(TaskResult).where(TaskResult.id == tid)).scalar_one_or_none()
        if not tr:
            tr = TaskResult(
                id=tid,
                task_name="import_meters_from_file",
                user_id=user_id,
                status=TaskStatus.PENDING,
                params={"file_name": file_name, "ext": file_type},
                progress={"current": 0, "total": 0, "success": 0, "failed": 0, "percent": 0},
                started_at=datetime.now(timezone.utc),
            )
            db.add(tr)
            db.commit()

        # === Lecture fichier ===
        raw = base64.b64decode(file_content_b64)
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = wb.active

        # Colonnes attendues
        RUS_COLS = {
            "id_code": "Идентификационный код",
            "address": "Адрес",
            "client_name": "Наименование объекта сети",
            "meter_type": "Тип прибора учета",
            "meter_number": "Номер ПУ",
        }
        # Désormais seule id_code est obligatoire
        REQUIRED = [RUS_COLS["id_code"]]

        headers_row2 = [str(c.value).strip() if c.value else None for c in sheet[2]]
        header_index = {h: i for i, h in enumerate(headers_row2) if h}
        missing = [c for c in REQUIRED if c not in header_index]
        if missing:
            raise ValueError(f"Colonnes manquantes: {missing}. Colonnes détectées: {headers_row2}")

        total_rows = max(0, sheet.max_row - 2)
        safe_progress(db, tid, 0, total_rows, 0, 0)

        success = failed = processed = 0
        errors: List[dict] = []
        seen_numbers = set()

        BATCH = 500
        PROGRESS_EVERY = 50
        buffer: List[dict] = []

        def get_val(row, colname):
            idx = header_index.get(colname)
            return row[idx] if idx is not None and idx < len(row) else None

        def flush():
            """Insert bulk avec ON CONFLICT DO NOTHING."""
            nonlocal buffer, success, failed
            if not buffer:
                return
            try:
                stmt = insert(Meter.__table__).values(buffer)
                stmt = stmt.on_conflict_do_nothing(index_elements=["meter_number"])
                db.execute(stmt)
                db.commit()
            except IntegrityError as e:
                logger.warning("Erreur intégrité, fallback unitaire: %s", e)
                db.rollback()
                for row in buffer:
                    try:
                        db.execute(
                            insert(Meter.__table__).values(row).on_conflict_do_nothing(index_elements=["meter_number"])
                        )
                        db.commit()
                    except Exception as ex:
                        db.rollback()
                        errors.append({"row": None, "meter_number": row.get("meter_number"), "error": str(ex)})
                        failed += 1
            finally:
                buffer.clear()

        # Parcours lignes 3..N
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            processed += 1
            try:
                meter_id_code = str(get_val(row, RUS_COLS["id_code"]) or "").strip()
                meter_number = str(get_val(row, RUS_COLS["meter_number"]) or "").strip()
                meter_type = str(get_val(row, RUS_COLS["meter_type"]) or "").strip()

                # Vérifie uniquement id_code
                if not meter_id_code:
                    failed += 1
                    errors.append({"row": row_idx, "error": "Champ obligatoire manquant: id_code"})
                    continue

                if meter_number and meter_number in seen_numbers:
                    failed += 1
                    errors.append({"row": row_idx, "meter_number": meter_number, "error": "Duplicate in file"})
                    continue
                if meter_number:
                    seen_numbers.add(meter_number)

                buffer.append({
                    "meter_id_code": meter_id_code,
                    "meter_number": meter_number or None,
                    "type": meter_type or None,
                    "location_address": str(get_val(row, RUS_COLS["address"]) or "").strip(),
                    "client_name": str(get_val(row, RUS_COLS["client_name"]) or "").strip(),
                    "status": "active",
                    "meter_metadata": {},
                })
                success += 1

                if len(buffer) >= BATCH:
                    flush()

            except Exception as e:
                failed += 1
                errors.append({"row": row_idx, "error": str(e)})

            if processed % PROGRESS_EVERY == 0:
                safe_progress(db, tid, processed, total_rows, success, failed)

        flush()
        safe_progress(db, tid, processed, total_rows, success, failed)

        result = {"file": file_name, "success": success, "failed": failed, "total": total_rows, "errors": errors[:200]}

        db.execute(
            update(TaskResult)
            .where(TaskResult.id == tid)
            .values(
                status=TaskStatus.COMPLETED,
                result=result,
                completed_at=datetime.now(timezone.utc),
                progress={"percent": 100, "current": processed, "total": total_rows, "success": success, "failed": failed},
            )
        )
        db.commit()

        return {"task_id": tid, "status": "completed", **result}

    except Exception as e:
        logger.exception(f"[{tid}] Import meters failed: {e}")
        db.rollback()
        db.execute(
            update(TaskResult)
            .where(TaskResult.id == tid)
            .values(status=TaskStatus.FAILED, error_message=str(e), completed_at=datetime.now(timezone.utc))
        )
        db.commit()
        raise
    finally:
        db.close()

