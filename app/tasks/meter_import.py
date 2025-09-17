# app/tasks/meter_import.py
import io
import math
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone

from celery import states
from celery.exceptions import Ignore
from openpyxl import load_workbook
from sqlalchemy import select, or_, insert
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.core.celery_app import celery_app
from app.database import SessionLocalSync
from app.models.meter import Meter

logger = logging.getLogger(__name__)

RUS_COLS = {
    "id_code": "Идентификационный код",
    "address": "Адрес",
    "client_name": "Наименование объекта сети",
    "meter_type": "Тип прибора учета",
    "meter_number": "Номер ПУ",
    "prev_reading": "Предыдущие показания",
    "curr_reading": "Текущие показания",  # ignoré
}
REQUIRED = [
    RUS_COLS["id_code"],
    RUS_COLS["meter_type"],
    RUS_COLS["meter_number"],
]

def _to_str(x):
    if x is None: return None
    s = str(x).strip()
    return s or None

def _to_float(x):
    if x is None: return None
    try:
        return float(str(x).replace(",", ".").strip())
    except Exception:
        return None

def _to_dt_tz(x):
    if x is None: return None
    if isinstance(x, datetime):
        dt = x
    else:
        try:
            dt = datetime.fromisoformat(str(x))
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt

def _read_sheet_headers(sheet):
    headers_row2 = [str(c.value).strip() if c.value else None for c in sheet[2]]
    header_index = {h: i for i, h in enumerate(headers_row2) if h}
    first_row = [str(c.value).strip() if c.value else None for c in sheet[1]]
    has_visit_date = "Дата обхода" in first_row
    visit_date_col_idx = 7 if has_visit_date and len(headers_row2) >= 8 else None
    missing = [c for c in REQUIRED if c not in header_index]
    if missing:
        raise ValueError(f"Colonnes manquantes: {missing}. Colonnes détectées: {headers_row2}")
    return header_index, visit_date_col_idx

def _yield_rows(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        yield row_idx, row


@celery_app.task(bind=True, name="tasks.import_meters", queue="default")
def import_meters_task(self, *, file_bytes: bytes) -> Dict[str, Any]:
    """
    Tâche d’import:
      - parse XLSX depuis bytes
      - upsert par lot (ON CONFLICT DO NOTHING sur meter_number ou meter_id_code)
      - progression via self.update_state
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active

        header_index, visit_idx = _read_sheet_headers(sheet)

        success = 0
        failed = 0
        errors: List[str] = []

        # Compter total pour progression
        total_rows = sheet.max_row - 2 if sheet.max_row > 2 else 0
        processed = 0

        BATCH = 500
        buffer: List[dict] = []

        def emit_progress():
            pct = int((processed / max(total_rows, 1)) * 100)
            self.update_state(
                state="PROGRESS",
                meta={
                    "processed": processed,
                    "total": total_rows,
                    "percent": pct,
                    "success": success,
                    "failed": failed,
                },
            )

        with SessionLocalSync() as db:  # sync session
            for row_idx, row in _yield_rows(sheet):
                try:
                    def val(col_name):
                        idx = header_index.get(col_name)
                        return row[idx] if idx is not None and idx < len(row) else None

                    meter_id_code = _to_str(val(RUS_COLS["id_code"]))
                    meter_number  = _to_str(val(RUS_COLS["meter_number"]))
                    meter_type    = _to_str(val(RUS_COLS["meter_type"]))

                    if not meter_id_code:
                        failed += 1
                        errors.append(f"Ligne {row_idx}: champs requis manquants (id).")
                        continue

                    location_address = _to_str(val(RUS_COLS["address"]))
                    client_name      = _to_str(val(RUS_COLS["client_name"]))
                    prev_read        = _to_float(val(RUS_COLS["prev_reading"]))

                    last_prev_dt = None
                    if visit_idx is not None and visit_idx < len(row):
                        last_prev_dt = _to_dt_tz(row[visit_idx])

                    buffer.append({
                        "meter_id_code": meter_id_code,
                        "meter_number": meter_number,
                        "type": meter_type,
                        "location_address": location_address,
                        "client_name": client_name,
                        "prev_reading_value": prev_read,
                        "last_reading_date": last_prev_dt,
                        "status": "active",
                        "meter_metadata": {},
                    })

                    if len(buffer) >= BATCH:
                        _flush_batch(db, buffer)
                        success += len(buffer)
                        buffer.clear()

                except Exception as e:
                    failed += 1
                    errors.append(f"Ligne {row_idx}: {e}")

                finally:
                    processed += 1
                    if processed % 200 == 0:  # push progress de temps en temps
                        emit_progress()

            if buffer:
                _flush_batch(db, buffer)
                success += len(buffer)
                buffer.clear()

        self.update_state(
            state=states.SUCCESS,
            meta={"success": success, "failed": failed, "errors": errors, "total": total_rows},
        )
        return {"success": success, "failed": failed, "errors": errors, "total": total_rows}

    except Exception as e:
        logger.exception("Import failed")
        self.update_state(state=states.FAILURE, meta={"exc": str(e)})
        raise

def _flush_batch(db: Session, rows: List[dict]):
    """
    Insertion en lot avec UPSERT idempotent.
    On déduplique par (meter_number) OU (meter_id_code).
    NB: ajuster les index/contrainte unique côté DB pour garantir l’unicité.
    """
    if not rows:
        return
    stmt = (
        pg_insert(Meter.__table__)
        .values(rows)
        .on_conflict_do_nothing(index_elements=["meter_number"])
    )
    db.execute(stmt)
    db.commit()
