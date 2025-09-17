from __future__ import annotations
from typing import Optional, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, desc
from datetime import date, datetime, timezone
import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.reading import Reading
from app.models.meter import Meter
from app.models.user import User
from app.services.storage_service import storage_service

logger = logging.getLogger(__name__)


class ExportService:
	def __init__(self, session: AsyncSession):
		self.session = session
		self.storage_service = storage_service

	def _draw_report_header(self, worksheet: Worksheet) -> None:
		"""Dessine les en-têtes avec cellules fusionnées"""
		# Première ligne avec cellules fusionnées
		worksheet.merge_cells("A1:E1")
		worksheet.merge_cells("F1:G1")
		worksheet.merge_cells("H1:H2")
		worksheet.merge_cells("I1:J1")
		worksheet.merge_cells("K1:L1")
		worksheet.merge_cells("M1:M2")
		worksheet.merge_cells("N1:N2")

		# Valeurs première ligne
		worksheet.cell(row=1, column=1).value = "Информация об элементе сети"
		worksheet.cell(row=1, column=6).value = "Показания"
		worksheet.cell(row=1, column=8).value = "Дата обхода"
		worksheet.cell(row=1, column=9).value = "Координаты"
		worksheet.cell(row=1, column=11).value = "Фотографии"
		worksheet.cell(row=1, column=13).value = "Исполнитель"
		worksheet.cell(row=1, column=14).value = "Комментарии"

		# Alignement et style première ligne
		for col in [1, 6, 8, 9, 11, 13, 14]:
			cell = worksheet.cell(row=1, column=col)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.font = Font(bold=True, size=11)

		# Deuxième ligne - sous-colonnes
		headers_row2 = [
			(1, "Идентификационный код"),
			(2, "Адрес"),
			(3, "Наименование объекта сети"),
			(4, "Тип прибора учета"),
			(5, "Номер ПУ"),
			(6, "Предыдущие показания"),
			(7, "Текущие показания"),
			(9, "Долгота"),
			(10, "Широта"),
			(11, "Показания"),
			(12, "Счетчик"),
		]

		for col, value in headers_row2:
			cell = worksheet.cell(row=2, column=col, value=value)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.font = Font(bold=True)

		# Largeurs de colonnes
		column_widths = {
			"A": 30, "B": 35, "C": 30, "D": 25, "E": 25,
			"F": 20, "G": 20, "H": 20, "I": 15, "J": 15,
			"K": 25, "L": 25, "M": 25, "N": 40
		}

		for col_letter, width in column_widths.items():
			worksheet.column_dimensions[col_letter].width = width

		# Hauteur des lignes d'en-tête
		worksheet.row_dimensions[1].height = 30
		worksheet.row_dimensions[2].height = 30

		# Bordures pour les en-têtes
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		for row in range(1, 3):
			for col in range(1, 15):
				worksheet.cell(row=row, column=col).border = thin_border

	async def export_readings(
			self,
			start_date: date,
			end_date: date,
			include_photos: bool = True,
			user_id: Optional[str] = None
	) -> io.BytesIO:
		"""
		Exporte les relevés en Excel avec la structure d'en-têtes définie.
		Retourne un BytesIO positionné au début.
		"""
		# 1) Récupération des données avec photos JSON
		async with self.session as db:
			query = (
				select(
					Reading.id.label("reading_id"),
					Reading.reading_value,
					Reading.reading_date,
					Reading.latitude.label("reading_latitude"),
					Reading.longitude.label("reading_longitude"),
					Reading.notes,
					Reading.photos,  # Ajout du champ photos JSON
					Meter.meter_number,
					Meter.type.label("meter_type"),
					Meter.location_address,
					Meter.client_name,
					Meter.prev_reading_value,
					Meter.meter_id_code,
					User.full_name.label("controller_name"),
				)
				.join(Meter, Reading.meter_id == Meter.id)
				.join(User, Reading.user_id == User.id)
				.where(
					and_(
						Reading.reading_date >= datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc),
						Reading.reading_date <= datetime.combine(end_date, datetime.max.time(), tzinfo=timezone.utc),
					)
				)
				.order_by(desc(Reading.reading_date))
			)
			if user_id:
				query = query.where(Reading.user_id == user_id)

			result = await db.execute(query)
			rows: List[tuple] = result.all()

		# 2) Création du workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Отчет по показаниям"

		# 3) Ajout des en-têtes personnalisés
		self._draw_report_header(ws)

		# 4) Styles pour les données
		date_style = NamedStyle(name="date_time_style")
		date_style.number_format = "dd.mm.yyyy hh:mm"
		num_style = NamedStyle(name="num_style")
		num_style.number_format = "#,##0"

		for st in (date_style, num_style):
			if st.name not in wb.named_styles:
				wb.add_named_style(st)

		# Bordure pour les cellules de données
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 5) Ajout des données (commence à la ligne 3)
		for row_idx, row in enumerate(rows, start=3):
			data = row._mapping

			# Colonne A: Идентификационный код
			ws.cell(row=row_idx, column=1, value=data.get("meter_id_code"))

			# Colonne B: Адрес
			ws.cell(row=row_idx, column=2, value=data.get("location_address"))

			# Colonne C: Наименование объекта сети (client_name)
			ws.cell(row=row_idx, column=3, value=data.get("client_name"))

			# Colonne D: Тип прибора учета
			ws.cell(row=row_idx, column=4, value=data.get("meter_type"))

			# Colonne E: Номер ПУ
			ws.cell(row=row_idx, column=5, value=data.get("meter_number"))

			# Colonne F: Предыдущие показания
			prev_value = data.get("prev_reading_value")
			if prev_value is not None:
				cell = ws.cell(row=row_idx, column=6, value=float(prev_value))
				cell.style = "num_style"

			# Colonne G: Текущие показания
			curr_value = data.get("reading_value")
			if curr_value is not None:
				cell = ws.cell(row=row_idx, column=7, value=float(curr_value))
				cell.style = "num_style"

			# Colonne H: Дата обхода
			reading_date = data.get("reading_date")
			if reading_date:
				if isinstance(reading_date, datetime):
					# Convert to timezone-naive datetime by removing tzinfo
					reading_date_naive = reading_date.replace(tzinfo=None)
					cell = ws.cell(row=row_idx, column=8, value=reading_date_naive)
				else:
					try:
						dt = datetime.fromisoformat(str(reading_date))
						# Ensure the parsed datetime is timezone-naive
						dt_naive = dt.replace(tzinfo=None)
						cell = ws.cell(row=row_idx, column=8, value=dt_naive)
					except Exception:
						cell = ws.cell(row=row_idx, column=8, value=str(reading_date))
				cell.style = "date_time_style"

			# Colonne I: Долгота
			longitude = data.get("reading_longitude")
			if longitude is not None:
				ws.cell(row=row_idx, column=9, value=f"{float(longitude):.6f}")

			# Colonne J: Широта
			latitude = data.get("reading_latitude")
			if latitude is not None:
				ws.cell(row=row_idx, column=10, value=f"{float(latitude):.6f}")

			# Colonnes K et L: Фотографии (liens courts)
			photos = data.get("photos", [])
			if photos and include_photos:
				# Colonne K: Photo des relevés
				if len(photos) > 0:
					self._add_photo_link(ws, row_idx, 11, photos[0], "Фото показаний")

				# Colonne L: Photo du compteur
				if len(photos) > 1:
					self._add_photo_link(ws, row_idx, 12, photos[1], "Фото счетчика")

			# Colonne M: Исполнитель
			ws.cell(row=row_idx, column=13, value=data.get("controller_name"))

			# Colonne N: Комментарии
			ws.cell(row=row_idx, column=14, value=data.get("notes", ""))

			# Appliquer les bordures à toutes les cellules de la ligne
			for col in range(1, 15):
				ws.cell(row=row_idx, column=col).border = thin_border

		# 6) Filtres automatiques (commence après les en-têtes fusionnés)
		if rows:
			ws.auto_filter.ref = f"A2:N{len(rows) + 2}"

		# Freeze panes après les deux lignes d'en-tête
		ws.freeze_panes = "A3"

		# 7) Ajout de l'onglet résumé
		self._add_summary_sheet(wb, rows, start_date, end_date)

		# 8) Sauvegarde dans BytesIO
		out = io.BytesIO()
		wb.save(out)
		out.seek(0)
		return out

	async def export_readings_all(
			self,
			include_photos: bool = True,
			user_id: Optional[str] = None
	) -> io.BytesIO:
		"""
		Exporte tous les relevés en Excel sans filtrer par date.
		Retourne un BytesIO positionné au début.
		"""
		# 1) Récupération de toutes les données avec photos JSON
		async with self.session as db:
			query = (
				select(
					Reading.id.label("reading_id"),
					Reading.reading_value,
					Reading.reading_date,
					Reading.latitude.label("reading_latitude"),
					Reading.longitude.label("reading_longitude"),
					Reading.notes,
					Reading.photos,
					Meter.meter_number,
					Meter.type.label("meter_type"),
					Meter.location_address,
					Meter.client_name,
					Meter.prev_reading_value,
					Meter.meter_id_code,
					User.full_name.label("controller_name"),
				)
				.join(Meter, Reading.meter_id == Meter.id)
				.join(User, Reading.user_id == User.id)
				.order_by(desc(Reading.reading_date))
			)
			if user_id:
				query = query.where(Reading.user_id == user_id)

			result = await db.execute(query)
			rows: List[tuple] = result.all()

		# 2) Création du workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Отчет по показаниям"

		# 3) Ajout des en-têtes personnalisés
		self._draw_report_header(ws)

		# 4) Styles pour les données
		date_style = NamedStyle(name="date_time_style")
		date_style.number_format = "dd.mm.yyyy hh:mm"
		num_style = NamedStyle(name="num_style")
		num_style.number_format = "#,##0"

		for st in (date_style, num_style):
			if st.name not in wb.named_styles:
				wb.add_named_style(st)

		# Bordure pour les cellules de données
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		# 5) Ajout des données (commence à la ligne 3)
		for row_idx, row in enumerate(rows, start=3):
			data = row._mapping

			# Colonne A: Идентификационный код
			ws.cell(row=row_idx, column=1, value=data.get("meter_id_code"))

			# Colonne B: Адрес
			ws.cell(row=row_idx, column=2, value=data.get("location_address"))

			# Colonne C: Наименование объекта сети (client_name)
			ws.cell(row=row_idx, column=3, value=data.get("client_name"))

			# Colonne D: Тип прибора учета
			ws.cell(row=row_idx, column=4, value=data.get("meter_type"))

			# Colonne E: Номер ПУ
			ws.cell(row=row_idx, column=5, value=data.get("meter_number"))

			# Colonne F: Предыдущие показания
			prev_value = data.get("prev_reading_value")
			if prev_value is not None:
				cell = ws.cell(row=row_idx, column=6, value=float(prev_value))
				cell.style = "num_style"

			# Colonne G: Текущие показания
			curr_value = data.get("reading_value")
			if curr_value is not None:
				cell = ws.cell(row=row_idx, column=7, value=float(curr_value))
				cell.style = "num_style"

			# Colonne H: Дата обхода
			reading_date = data.get("reading_date")
			if reading_date:
				if isinstance(reading_date, datetime):
					# Convert to timezone-naive datetime by removing tzinfo
					reading_date_naive = reading_date.replace(tzinfo=None)
					cell = ws.cell(row=row_idx, column=8, value=reading_date_naive)
				else:
					try:
						dt = datetime.fromisoformat(str(reading_date))
						# Ensure the parsed datetime is timezone-naive
						dt_naive = dt.replace(tzinfo=None)
						cell = ws.cell(row=row_idx, column=8, value=dt_naive)
					except Exception:
						cell = ws.cell(row=row_idx, column=8, value=str(reading_date))
				cell.style = "date_time_style"

			# Colonne I: Долгота
			longitude = data.get("reading_longitude")
			if longitude is not None:
				ws.cell(row=row_idx, column=9, value=f"{float(longitude):.6f}")

			# Colonne J: Широта
			latitude = data.get("reading_latitude")
			if latitude is not None:
				ws.cell(row=row_idx, column=10, value=f"{float(latitude):.6f}")

			# Colonnes K et L: Фотографии (liens courts)
			photos = data.get("photos", [])
			if photos and include_photos:
				# Colonne K: Photo des relevés
				if len(photos) > 0:
					self._add_photo_link(ws, row_idx, 11, photos[0], "Фото показаний")

				# Colonne L: Photo du compteur
				if len(photos) > 1:
					self._add_photo_link(ws, row_idx, 12, photos[1], "Фото счетчика")

			# Colonne M: Исполнитель
			ws.cell(row=row_idx, column=13, value=data.get("controller_name"))

			# Colonne N: Комментарии
			ws.cell(row=row_idx, column=14, value=data.get("notes", ""))

			# Appliquer les bordures à toutes les cellules de la ligne
			for col in range(1, 15):
				ws.cell(row=row_idx, column=col).border = thin_border

		# 6) Filtres automatiques (commence après les en-têtes fusionnés)
		if rows:
			ws.auto_filter.ref = f"A2:N{len(rows) + 2}"

		# Freeze panes après les deux lignes d'en-tête
		ws.freeze_panes = "A3"

		# 7) Ajout de l'onglet résumé (sans période)
		self._add_summary_sheet_all(wb, rows)

		# 8) Sauvegarde dans BytesIO
		out = io.BytesIO()
		wb.save(out)
		out.seek(0)
		return out

	def _add_summary_sheet_all(self, wb: Workbook, rows: List[tuple]):
		"""Crée un onglet 'Сводка' avec les statistiques, sans période spécifique."""
		ws = wb.create_sheet("Сводка")
		ws.column_dimensions["A"].width = 35
		ws.column_dimensions["B"].width = 30

		# Style pour le titre
		title_font = Font(bold=True, size=14)
		header_font = Font(bold=True, size=11)

		# Titre
		title = ws.cell(row=1, column=1, value="Сводка по экспорту (tous les relevés)")
		title.font = title_font

		# Total des relevés
		total_readings = len(rows)
		ws.cell(row=3, column=1, value="Всего показаний")
		ws.cell(row=3, column=2, value=total_readings)

		# Compteurs uniques et contrôleurs
		meter_numbers = set()
		controllers = set()
		meter_types_count = {}

		for r in rows:
			m = r._mapping
			if m.get("meter_number", "Не указан"):
				meter_numbers.add(m["meter_number"])
			if m.get("controller_name"):
				controllers.add(m["controller_name"])
			mt = m.get("meter_type", "Не указан")
			meter_types_count[mt] = meter_types_count.get(mt, 0) + 1

		ws.cell(row=4, column=1, value="Уникальных приборов учета")
		ws.cell(row=4, column=2, value=len(meter_numbers))

		ws.cell(row=5, column=1, value="Контролеров")
		ws.cell(row=5, column=2, value=len(controllers))

		# Statistiques par type
		ws.cell(row=7, column=1, value="Показания по типам приборов")
		ws.cell(row=7, column=1).font = header_font

		row_cursor = 8
		for mt, count in sorted(meter_types_count.items(), key=lambda x: x[0] or ""):
			ws.cell(row=row_cursor, column=1, value=mt)
			ws.cell(row=row_cursor, column=2, value=count)
			row_cursor += 1

		# Bordures pour la section de résumé
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		for row in range(3, row_cursor):
			for col in range(1, 3):
				ws.cell(row=row, column=col).border = thin_border

	def _add_photo_link(self, ws: Worksheet, row: int, col: int, url: str, display_text: str):
		"""
		Ajoute un lien hypertexte court dans une cellule.

		Args:
			ws: La feuille de calcul
			row: Numéro de ligne
			col: Numéro de colonne
			url: URL complète de la photo
			display_text: Texte affiché pour le lien
		"""
		cell = ws.cell(row=row, column=col)
		cell.value = display_text
		cell.hyperlink = url
		cell.font = Font(color="0563C1", underline="single")  # Bleu avec soulignement
		cell.alignment = Alignment(horizontal="center", vertical="center")

	def _add_summary_sheet(self, wb: Workbook, rows: List[tuple], start_date: date, end_date: date):
		"""Crée un onglet 'Сводка' avec les statistiques."""
		ws = wb.create_sheet("Сводка")
		ws.column_dimensions["A"].width = 35
		ws.column_dimensions["B"].width = 30

		# Style pour le titre
		title_font = Font(bold=True, size=14)
		header_font = Font(bold=True, size=11)

		# Titre
		title = ws.cell(row=1, column=1, value="Сводка по экспорту")
		title.font = title_font

		# Période
		ws.cell(row=3, column=1, value="Период")
		ws.cell(row=3, column=2, value=f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")

		# Total des relevés
		total_readings = len(rows)
		ws.cell(row=4, column=1, value="Всего показаний")
		ws.cell(row=4, column=2, value=total_readings)

		# Compteurs uniques et contrôleurs
		meter_numbers = set()
		controllers = set()
		meter_types_count = {}

		for r in rows:
			m = r._mapping
			if m.get("meter_number"):
				meter_numbers.add(m["meter_number"])
			if m.get("controller_name"):
				controllers.add(m["controller_name"])
			mt = m.get("meter_type", "Не указан")
			meter_types_count[mt] = meter_types_count.get(mt, 0) + 1

		ws.cell(row=5, column=1, value="Уникальных приборов учета")
		ws.cell(row=5, column=2, value=len(meter_numbers))

		ws.cell(row=6, column=1, value="Контролеров")
		ws.cell(row=6, column=2, value=len(controllers))

		# Statistiques par type
		ws.cell(row=8, column=1, value="Показания по типам приборов")
		ws.cell(row=8, column=1).font = header_font

		row_cursor = 9
		for mt, count in sorted(meter_types_count.items(), key=lambda x: x[0] or ""):
			ws.cell(row=row_cursor, column=1, value=mt)
			ws.cell(row=row_cursor, column=2, value=count)
			row_cursor += 1

		# Bordures pour la section de résumé
		thin_border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)

		for row in range(3, row_cursor):
			for col in range(1, 3):
				ws.cell(row=row, column=col).border = thin_border