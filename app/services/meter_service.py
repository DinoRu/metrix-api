# app/services/meter_service.py

from typing import Dict, Any, List
from fastapi import UploadFile, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_
from datetime import datetime, timezone
import logging, io

from openpyxl import load_workbook

from app.models.meter import Meter
from app.schemas.meter import MeterResponse

logger = logging.getLogger(__name__)

# Colonnes réelles (ligne 2)
RUS_COLS = {
    "id_code": "Идентификационный код",
    "address": "Адрес",
    "client_name": "Наименование объекта сети",
    "meter_type": "Тип прибора учета",
    "meter_number": "Номер ПУ",
    "prev_reading": "Предыдущие показания",   # → Meter.prev_reading_value
    "curr_reading": "Текущие показания",      # ← ignoré (vide à l’import)
}
REQUIRED = [
    RUS_COLS["id_code"],
    RUS_COLS["meter_type"],
    RUS_COLS["meter_number"],
]

def _to_str(x):
    if x is None:
        return None
    s = str(x).strip()
    return s or None

def _to_float(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", ".").strip())
    except Exception:
        return None

def _to_dt_tz(x):
    """Normalise une date en datetime tz-aware (UTC)."""
    if x is None:
        return None
    if isinstance(x, datetime):
        dt = x
    else:
        try:
            dt = datetime.fromisoformat(str(x))
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


class MeterService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def import_from_file(self, file: UploadFile, *, current_user=None) -> Dict[str, Any]:
        """
        Import des compteurs depuis un fichier XLSX.
        - Ignore 'Текущие показания' (les nouveaux relevés seront saisis plus tard).
        - 'Предыдущие показания' → prev_reading_value sur Meter.
        - 'Дата обхода' (si présente) → last_reading_date.
        """
        try:
            if not file.filename.lower().endswith(".xlsx"):
                raise ValueError("Format non supporté : fournir un fichier .xlsx")

            content = await file.read()
            wb = load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active

            # Ligne 1 = groupes fusionnés, Ligne 2 = en-têtes réels
            headers_row2 = [str(c.value).strip() if c.value else None for c in sheet[2]]
            header_index = {h: i for i, h in enumerate(headers_row2) if h}

            # Détection optionnelle de "Дата обхода" (souvent en 8e colonne)
            first_row = [str(c.value).strip() if c.value else None for c in sheet[1]]
            has_visit_date = "Дата обхода" in first_row
            visit_date_col_idx = 7 if has_visit_date and len(headers_row2) >= 8 else None  # 0-based

            # Vérifs colonnes requises
            missing = [c for c in REQUIRED if c not in header_index]
            if missing:
                raise ValueError(
                    f"Colonnes manquantes: {missing}. Colonnes détectées (ligne 2): {headers_row2}"
                )

            success, failed = 0, 0
            errors: List[str] = []
            meters: List[MeterResponse] = []

            async with self.session as db:
                # Données à partir de la ligne 3
                for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    try:
                        def val(col_name):
                            idx = header_index.get(col_name)
                            return row[idx] if idx is not None and idx < len(row) else None

                        meter_id_code = _to_str(val(RUS_COLS["id_code"]))
                        meter_number  = _to_str(val(RUS_COLS["meter_number"]))
                        meter_type    = _to_str(val(RUS_COLS["meter_type"]))

                        if not meter_id_code or not meter_number or not meter_type:
                            failed += 1
                            errors.append(f"Ligne {row_idx}: champs requis manquants (id/num/type).")
                            continue

                        # Doublon si même meter_number OU même meter_id_code
                        existing_q = await db.execute(
                            select(Meter).where(
                                or_(Meter.meter_number == meter_number,
                                    Meter.meter_id_code == meter_id_code)
                            )
                        )
                        if existing_q.scalar_one_or_none():
                            failed += 1
                            errors.append(
                                f"Ligne {row_idx}: compteur {meter_number}/{meter_id_code} existe déjà."
                            )
                            continue

                        location_address = _to_str(val(RUS_COLS["address"]))
                        client_name      = _to_str(val(RUS_COLS["client_name"]))
                        prev_read        = _to_float(val(RUS_COLS["prev_reading"]))   # ← stocké sur Meter
                        # curr_read        = _to_float(val(RUS_COLS["curr_reading"])) # ← ignoré à l'import

                        # Date de passage si fournie (considérée comme date du relevé précédent)
                        last_prev_dt = None
                        if visit_date_col_idx is not None and visit_date_col_idx < len(row):
                            last_prev_dt = _to_dt_tz(row[visit_date_col_idx])

                        meter = Meter(
                            meter_id_code=meter_id_code,
                            meter_number=meter_number,
                            type=meter_type,
                            location_address=location_address,
                            client_name=client_name,
                            prev_reading_value=prev_read,
                            last_reading_date=last_prev_dt,
                            status="active",
                            meter_metadata={},  # minimal
                        )
                        db.add(meter)
                        await db.flush()

                        meters.append(MeterResponse.model_validate(meter))
                        success += 1

                    except Exception as e:
                        failed += 1
                        errors.append(f"Ligne {row_idx}: {str(e)}")

                if success > 0:
                    await db.commit()

            logger.info(f"Import terminé: {success} succès, {failed} échecs")
            return {"success": success, "failed": failed, "errors": errors, "meters": meters}

        except Exception as e:
            logger.error(f"Échec import: {str(e)}")
            raise HTTPException(status_code=400, detail=str(e))
